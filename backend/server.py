"""MesureChâssis backend - FastAPI + MongoDB."""
from __future__ import annotations

import io
import logging
import math
import os
import uuid
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Annotated, Any, List, Optional

import httpx
import jwt
from dotenv import load_dotenv
from fastapi import APIRouter, Depends, FastAPI, HTTPException, status
from fastapi.responses import StreamingResponse, Response
from motor.motor_asyncio import AsyncIOMotorClient
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from passlib.context import CryptContext
from pydantic import BaseModel, EmailStr, Field, field_validator
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer, Table,
                                TableStyle)
from starlette.middleware.cors import CORSMiddleware

# --- Setup ---------------------------------------------------------------
ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s")
logger = logging.getLogger("mesurechassis")

MONGO_URL = os.environ["MONGO_URL"]
DB_NAME = os.environ["DB_NAME"]
JWT_SECRET = os.environ.get("JWT_SECRET", "change-me-mesurechassis-secret-key-dev-only")
JWT_ALGO = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60 * 24 * 7  # 7 days

client = AsyncIOMotorClient(MONGO_URL)
db = client[DB_NAME]

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

app = FastAPI(title="MesureChâssis API")
api = APIRouter(prefix="/api")

VALID_ROLES = {"admin", "commercial", "technician"}
VALID_STATUSES = {"devis_a_faire", "technique_a_valider", "en_commande", "en_fabrication", "cloture"}
CONVERTED_STATUSES = {"en_commande", "en_fabrication", "cloture"}
VALID_BLOCK_TYPES = {"standard", "coulissant", "porte", "trapeze"}
VALID_WALL_TYPES = {"ite", "iti", "brique_parement", "crepi_simple"}


# --- Models --------------------------------------------------------------
class UserPublic(BaseModel):
    id: str
    name: str
    email: EmailStr
    role: str
    company_id: str


class UserCreate(BaseModel):
    name: str
    email: EmailStr
    password: str
    role: str = "technician"
    company_id: str = "default"


class LoginRequest(BaseModel):
    email: EmailStr
    password: str


class TokenResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: UserPublic


class ChantierCreate(BaseModel):
    client_name: Optional[str] = None  # legacy fallback (kept for back-compat)
    first_name: Optional[str] = None   # Prénom
    last_name: Optional[str] = None    # Nom
    address: str
    postal_code: Optional[str] = None  # Code postal (string to keep leading zeros)
    city: Optional[str] = None         # Ville
    status: str = "devis_a_faire"
    assigned_to: Optional[str] = None
    appointment_at: Optional[str] = None  # ISO datetime
    notes: Optional[str] = None


class ChantierUpdate(BaseModel):
    client_name: Optional[str] = None
    first_name: Optional[str] = None
    last_name: Optional[str] = None
    address: Optional[str] = None
    postal_code: Optional[str] = None
    city: Optional[str] = None
    status: Optional[str] = None
    assigned_to: Optional[str] = None
    appointment_at: Optional[str] = None
    notes: Optional[str] = None


class Chantier(BaseModel):
    id: str
    client_name: str
    first_name: Optional[str] = None
    last_name: Optional[str] = None
    address: str
    postal_code: Optional[str] = None
    city: Optional[str] = None
    status: str
    created_by: str
    assigned_to: Optional[str] = None
    appointment_at: Optional[str] = None
    notes: Optional[str] = None
    company_id: str = "default"
    client_signature: Optional[str] = None
    signed_at: Optional[str] = None
    created_at: str


class CompanyProfile(BaseModel):
    company_id: str
    name: Optional[str] = None
    artisan_mode: bool = False


class CompanyProfileUpdate(BaseModel):
    name: Optional[str] = None
    artisan_mode: Optional[bool] = None


class SignatureIn(BaseModel):
    signature: str  # base64 data URL or raw base64


class MesureCreate(BaseModel):
    chantier_id: str
    block_type: str
    label: str
    # Legacy schema fields (optional, kept for backward compat)
    width_top: Optional[float] = None
    width_middle: Optional[float] = None
    width_bottom: Optional[float] = None
    height_left: Optional[float] = None
    height_middle: Optional[float] = None
    height_right: Optional[float] = None
    diag_1: Optional[float] = None
    diag_2: Optional[float] = None
    height_quarter_left: Optional[float] = None
    height_quarter_right: Optional[float] = None
    height_small: Optional[float] = None
    height_large: Optional[float] = None
    width_small: Optional[float] = None
    width_intermediate: Optional[float] = None
    # NEW — Baie brute (raw masonry bay) — Step 2
    bay_height: Optional[float] = None        # Hauteur (mm)
    bay_width: Optional[float] = None         # Largeur (mm)
    bay_diagonal: Optional[float] = None      # legacy single diagonal
    bay_diagonal_1: Optional[float] = None    # iter6 — Diagonale 1
    bay_diagonal_2: Optional[float] = None    # iter6 — Diagonale 2
    diag_1_verified: Optional[bool] = None    # validated by user (auto or manual)
    diag_2_verified: Optional[bool] = None
    floor_reserve: Optional[float] = None     # Réserve Sol Fini (mm) — only "porte"
    # NEW — Conception maçonnerie & isolation (indicatif) — Step 3
    bloc_thickness: Optional[float] = None    # Épaisseur Bloc Béton (mm)
    wall_type: Optional[str] = None           # "ite" | "iti" | "crepi_simple"
    insulation_thickness: Optional[float] = None  # ITE/ITI — Épaisseur Isolant
    finish_outer: Optional[float] = None      # ITE crépi / CRÉPI ext
    finish_inner: Optional[float] = None      # ITI plâtre / CRÉPI int
    options: dict = Field(default_factory=dict)
    photo_url: Optional[str] = None  # base64 data URL

    @field_validator("wall_type")
    @classmethod
    def _validate_wall_type(cls, v: Optional[str]) -> Optional[str]:
        if v is None:
            return v
        if v not in VALID_WALL_TYPES:
            raise ValueError(
                f"wall_type must be one of {sorted(VALID_WALL_TYPES)}"
            )
        return v


class Mesure(MesureCreate):
    id: str
    created_at: str
    alerts: List[str] = Field(default_factory=list)
    slope_angle_deg: Optional[float] = None


class FeedbackCreate(BaseModel):
    page_context: str
    user_comment: str
    screenshot_data: Optional[str] = None  # base64
    encoded_data_snapshot: dict = Field(default_factory=dict)


class Feedback(BaseModel):
    id: str
    user_id: str
    user_email: str
    page_context: str
    user_comment: str
    screenshot_data: Optional[str] = None
    encoded_data_snapshot: dict
    company_id: str = "default"
    created_at: str


# --- Helpers -------------------------------------------------------------
def hash_password(p: str) -> str:
    return pwd_context.hash(p)


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(user_id: str, role: str) -> str:
    expire = datetime.now(timezone.utc) + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    return jwt.encode({"sub": user_id, "role": role, "exp": expire},
                      JWT_SECRET, algorithm=JWT_ALGO)


def user_to_public(doc: dict) -> UserPublic:
    return UserPublic(id=doc["id"], name=doc["name"], email=doc["email"],
                      role=doc["role"], company_id=doc.get("company_id", "default"))


async def get_current_user(authorization: Annotated[Optional[str], Depends(lambda: None)] = None,
                           **kwargs) -> dict:
    raise NotImplementedError  # replaced below by header dep


# Real dep using Header
from fastapi import Header


async def auth_user(authorization: Optional[str] = Header(None)) -> dict:
    if not authorization or not authorization.lower().startswith("bearer "):
        raise HTTPException(status_code=401, detail="Missing token")
    token = authorization.split(" ", 1)[1].strip()
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGO])
    except jwt.PyJWTError:
        raise HTTPException(status_code=401, detail="Invalid token")
    user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=401, detail="User not found")
    # Attach company artisan_mode (best-effort)
    company_doc = await db.companies.find_one(
        {"company_id": user.get("company_id", "default")}, {"_id": 0})
    user["artisan_mode"] = bool(company_doc and company_doc.get("artisan_mode"))
    return user


def require_admin(user: dict = Depends(auth_user)) -> dict:
    if user["role"] != "admin" and not user.get("artisan_mode"):
        raise HTTPException(status_code=403, detail="Admin only")
    return user


def require_roles(roles: List[str]):
    def _dep(user: dict = Depends(auth_user)) -> dict:
        # Mode Artisan Unique bypasses all role restrictions
        if user.get("artisan_mode"):
            return user
        if user["role"] not in roles:
            raise HTTPException(
                status_code=403,
                detail=f"Réservé aux rôles : {', '.join(roles)}",
            )
        return user
    return _dep


# --- Expo Push ----------------------------------------------------------
EXPO_PUSH_URL = "https://exp.host/--/api/v2/push/send"


async def send_push_to_user(user_id: str, title: str, body: str,
                             data: Optional[dict] = None) -> None:
    """Best-effort push: never raises. Skips if user has no token."""
    u = await db.users.find_one({"id": user_id}, {"_id": 0, "push_token": 1})
    if not u or not u.get("push_token"):
        return
    payload = {
        "to": u["push_token"],
        "title": title,
        "body": body,
        "sound": "default",
        "data": data or {},
    }
    try:
        async with httpx.AsyncClient(timeout=10) as cli:
            r = await cli.post(EXPO_PUSH_URL, json=payload,
                                headers={"Accept-Encoding": "gzip, deflate",
                                          "Accept": "application/json"})
            if r.status_code >= 400:
                logger.warning("Push failed [%s]: %s", r.status_code, r.text[:200])
    except Exception as exc:
        logger.warning("Push send error: %s", exc)


class PushTokenIn(BaseModel):
    push_token: Optional[str] = None


# --- Calculation logic ---------------------------------------------------
def compute_alerts(m: MesureCreate) -> tuple[list[str], Optional[float]]:
    alerts: list[str] = []
    slope: Optional[float] = None
    bt = m.block_type
    if bt == "standard":
        widths = [v for v in (m.width_top, m.width_middle, m.width_bottom) if v is not None]
        heights = [v for v in (m.height_left, m.height_middle, m.height_right) if v is not None]
        if widths and (max(widths) - min(widths)) > 5:
            alerts.append("⚠️ Faux-aplomb détecté (largeurs)")
        if heights and (max(heights) - min(heights)) > 5:
            alerts.append("⚠️ Faux-aplomb détecté (hauteurs)")
        if m.diag_1 is not None and m.diag_2 is not None and abs(m.diag_1 - m.diag_2) > 5:
            alerts.append("⚠️ Hors-équerre")
    elif bt == "coulissant":
        widths = [m.width_top, m.width_middle, m.width_bottom]
        heights = [m.height_left, m.height_quarter_left, m.height_middle,
                   m.height_quarter_right, m.height_right]
        if any(v is None for v in widths):
            alerts.append("ℹ️ 3 largeurs requises (haut/milieu/bas)")
        if any(v is None for v in heights):
            alerts.append("ℹ️ 5 hauteurs requises pour détecter la flèche du linteau")
        valid_h = [v for v in heights if v is not None]
        if len(valid_h) >= 3 and (max(valid_h) - min(valid_h)) > 5:
            alerts.append("⚠️ Flèche du linteau détectée")
    elif bt == "trapeze":
        if (m.width_small is not None and m.width_intermediate is not None
                and m.height_small is not None and m.height_large is not None):
            dw = abs(m.width_intermediate - m.width_small)
            dh = abs(m.height_large - m.height_small)
            if dw > 0:
                slope = round(math.degrees(math.atan(dh / dw)), 2)
    return alerts, slope


# --- Auth routes ---------------------------------------------------------
@api.post("/auth/register", response_model=TokenResponse)
async def register(payload: UserCreate):
    if payload.role not in VALID_ROLES:
        raise HTTPException(400, "Invalid role")
    existing = await db.users.find_one({"email": payload.email.lower()})
    if existing:
        raise HTTPException(400, "Email already registered")
    user_doc = {
        "id": str(uuid.uuid4()),
        "name": payload.name,
        "email": payload.email.lower(),
        "role": payload.role,
        "company_id": payload.company_id,
        "hashed_password": hash_password(payload.password),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.users.insert_one(user_doc)
    token = create_access_token(user_doc["id"], user_doc["role"])
    return TokenResponse(access_token=token, user=user_to_public(user_doc))


@api.post("/auth/login", response_model=TokenResponse)
async def login(payload: LoginRequest):
    user = await db.users.find_one({"email": payload.email.lower()})
    if not user or not verify_password(payload.password, user["hashed_password"]):
        raise HTTPException(401, "Email ou mot de passe incorrect")
    token = create_access_token(user["id"], user["role"])
    return TokenResponse(access_token=token, user=user_to_public(user))


@api.get("/auth/me", response_model=UserPublic)
async def me(user=Depends(auth_user)):
    return user_to_public(user)


@api.get("/users", response_model=List[UserPublic])
async def list_users(user=Depends(auth_user)):
    docs = await db.users.find({"company_id": user.get("company_id", "default")},
                                {"_id": 0, "hashed_password": 0}).to_list(500)
    return [user_to_public(d) for d in docs]


@api.post("/auth/push-token")
async def set_push_token(payload: PushTokenIn, user=Depends(auth_user)):
    await db.users.update_one(
        {"id": user["id"]}, {"$set": {"push_token": payload.push_token}})
    return {"ok": True}


# --- Chantiers routes ----------------------------------------------------
@api.post("/chantiers", response_model=Chantier)
async def create_chantier(payload: ChantierCreate,
                           user=Depends(require_roles(["admin", "commercial"]))):
    if payload.status not in VALID_STATUSES:
        raise HTTPException(400, "Invalid status")
    # Backwards-compat: build client_name from first/last if missing
    client_name = payload.client_name
    if not client_name:
        parts = [p for p in [payload.last_name, payload.first_name] if p]
        client_name = " ".join(parts).strip() or "Sans nom"
    doc = {
        "id": str(uuid.uuid4()),
        "client_name": client_name,
        "first_name": payload.first_name,
        "last_name": payload.last_name,
        "address": payload.address,
        "postal_code": payload.postal_code,
        "city": payload.city,
        "status": payload.status,
        "created_by": user["id"],
        "assigned_to": payload.assigned_to,
        "appointment_at": payload.appointment_at,
        "notes": payload.notes,
        "company_id": user.get("company_id", "default"),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.chantiers.insert_one(doc)
    doc.pop("_id", None)
    if payload.assigned_to:
        await send_push_to_user(
            payload.assigned_to,
            "📌 Nouveau chantier assigné",
            f"{client_name} — Prise de rendez-vous à faire",
            {"type": "chantier_assigned", "chantier_id": doc["id"]},
        )
    return Chantier(**doc)


@api.get("/chantiers", response_model=List[Chantier])
async def list_chantiers(status_filter: Optional[str] = None, q: Optional[str] = None,
                          user=Depends(auth_user)):
    query: dict = {"company_id": user.get("company_id", "default")}
    if status_filter and status_filter in VALID_STATUSES:
        query["status"] = status_filter
    if q:
        import re as _re
        safe = _re.escape(q.strip())
        query["$or"] = [
            {"client_name": {"$regex": safe, "$options": "i"}},
            {"address": {"$regex": safe, "$options": "i"}},
        ]
    docs = await db.chantiers.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return [Chantier(**d) for d in docs]


@api.get("/chantiers/{chantier_id}", response_model=Chantier)
async def get_chantier(chantier_id: str, user=Depends(auth_user)):
    doc = await db.chantiers.find_one(
        {"id": chantier_id, "company_id": user.get("company_id", "default")},
        {"_id": 0})
    if not doc:
        raise HTTPException(404, "Chantier introuvable")
    return Chantier(**doc)


@api.patch("/chantiers/{chantier_id}", response_model=Chantier)
async def update_chantier(chantier_id: str, payload: ChantierUpdate,
                           user=Depends(require_roles(["admin", "commercial"]))):
    company = user.get("company_id", "default")
    existing = await db.chantiers.find_one({"id": chantier_id, "company_id": company})
    if not existing:
        raise HTTPException(404, "Chantier introuvable")
    update = {k: v for k, v in payload.model_dump().items() if v is not None}
    if "status" in update and update["status"] not in VALID_STATUSES:
        raise HTTPException(400, "Invalid status")
    if update:
        await db.chantiers.update_one({"id": chantier_id, "company_id": company}, {"$set": update})
    doc = await db.chantiers.find_one({"id": chantier_id, "company_id": company}, {"_id": 0})

    # Push notification on assignment change
    new_assignee = update.get("assigned_to")
    if new_assignee and new_assignee != existing.get("assigned_to"):
        await send_push_to_user(
            new_assignee,
            "Nouveau chantier affecté",
            f"{doc['client_name']} — {doc['address']}",
            {"type": "chantier_assigned", "chantier_id": chantier_id},
        )
    return Chantier(**doc)


@api.delete("/chantiers/{chantier_id}")
async def delete_chantier(chantier_id: str,
                           user=Depends(require_roles(["admin", "commercial"]))):
    company = user.get("company_id", "default")
    res = await db.chantiers.delete_one({"id": chantier_id, "company_id": company})
    if res.deleted_count:
        await db.mesures.delete_many({"chantier_id": chantier_id})
    return {"ok": True}


# --- Mesures routes ------------------------------------------------------
async def _check_chantier_access(chantier_id: str, user: dict) -> dict:
    chantier = await db.chantiers.find_one(
        {"id": chantier_id, "company_id": user.get("company_id", "default")})
    if not chantier:
        raise HTTPException(404, "Chantier introuvable")
    return chantier


@api.post("/mesures", response_model=Mesure)
async def create_mesure(payload: MesureCreate, user=Depends(auth_user)):
    if payload.block_type not in VALID_BLOCK_TYPES:
        raise HTTPException(400, "Invalid block_type")
    await _check_chantier_access(payload.chantier_id, user)
    alerts, slope = compute_alerts(payload)
    doc = payload.model_dump()
    doc.update({
        "id": str(uuid.uuid4()),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "alerts": alerts,
        "slope_angle_deg": slope,
    })
    await db.mesures.insert_one(doc)
    doc.pop("_id", None)
    return Mesure(**doc)


@api.get("/chantiers/{chantier_id}/mesures", response_model=List[Mesure])
async def list_mesures(chantier_id: str, user=Depends(auth_user)):
    await _check_chantier_access(chantier_id, user)
    docs = await db.mesures.find({"chantier_id": chantier_id}, {"_id": 0}) \
        .sort("created_at", 1).to_list(500)
    return [Mesure(**d) for d in docs]


@api.delete("/mesures/{mesure_id}")
async def delete_mesure(mesure_id: str, user=Depends(auth_user)):
    mesure = await db.mesures.find_one({"id": mesure_id})
    if mesure:
        await _check_chantier_access(mesure["chantier_id"], user)
        await db.mesures.delete_one({"id": mesure_id})
    return {"ok": True}


# --- Feedback routes -----------------------------------------------------
@api.post("/feedbacks", response_model=Feedback)
async def create_feedback(payload: FeedbackCreate, user=Depends(auth_user)):
    doc = {
        "id": str(uuid.uuid4()),
        "user_id": user["id"],
        "user_email": user["email"],
        "page_context": payload.page_context,
        "user_comment": payload.user_comment,
        "screenshot_data": payload.screenshot_data,
        "encoded_data_snapshot": payload.encoded_data_snapshot,
        "company_id": user.get("company_id", "default"),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.feedbacks.insert_one(doc)
    doc.pop("_id", None)
    return Feedback(**doc)


@api.get("/feedbacks", response_model=List[Feedback])
async def list_feedbacks(user=Depends(require_admin)):
    docs = await db.feedbacks.find(
        {"company_id": user.get("company_id", "default")},
        {"_id": 0}).sort("created_at", -1).to_list(500)
    return [Feedback(**d) for d in docs]


@api.delete("/feedbacks/{feedback_id}")
async def delete_feedback(feedback_id: str, user=Depends(require_admin)):
    await db.feedbacks.delete_one(
        {"id": feedback_id, "company_id": user.get("company_id", "default")})
    return {"ok": True}


# --- Company profile (Mode Artisan Unique) -----------------------------
@api.get("/company/profile", response_model=CompanyProfile)
async def get_company_profile(user=Depends(auth_user)):
    company_id = user.get("company_id", "default")
    doc = await db.companies.find_one({"company_id": company_id}, {"_id": 0})
    if not doc:
        doc = {"company_id": company_id, "name": company_id, "artisan_mode": False}
    return CompanyProfile(
        company_id=doc.get("company_id", company_id),
        name=doc.get("name") or company_id,
        artisan_mode=bool(doc.get("artisan_mode", False)),
    )


@api.patch("/company/profile", response_model=CompanyProfile)
async def update_company_profile(payload: CompanyProfileUpdate,
                                  user=Depends(require_admin)):
    company_id = user.get("company_id", "default")
    update = {k: v for k, v in payload.model_dump().items() if v is not None}
    if not update:
        # Read-through if nothing changed
        return await get_company_profile(user)
    update["company_id"] = company_id
    await db.companies.update_one(
        {"company_id": company_id},
        {"$set": update},
        upsert=True,
    )
    doc = await db.companies.find_one({"company_id": company_id}, {"_id": 0})
    return CompanyProfile(
        company_id=doc.get("company_id", company_id),
        name=doc.get("name") or company_id,
        artisan_mode=bool(doc.get("artisan_mode", False)),
    )


# --- Stats route ---------------------------------------------------------
@api.get("/stats/company")
async def stats_company(user=Depends(require_admin)):
    company = user.get("company_id", "default")
    pipe_status = [
        {"$match": {"company_id": company}},
        {"$group": {"_id": "$status", "count": {"$sum": 1}}},
    ]
    by_status: dict[str, int] = {s: 0 for s in VALID_STATUSES}
    async for d in db.chantiers.aggregate(pipe_status):
        if d["_id"]:
            by_status[d["_id"]] = d["count"]
    total = sum(by_status.values())
    closure_rate = round((by_status["cloture"] / total) * 100, 1) if total else 0.0

    company_chantier_ids = await db.chantiers.distinct("id", {"company_id": company})
    chantier_to_tech: dict[str, Optional[str]] = {}
    async for c in db.chantiers.find(
            {"company_id": company}, {"_id": 0, "id": 1, "assigned_to": 1}):
        chantier_to_tech[c["id"]] = c.get("assigned_to")

    by_tech: dict[str, dict] = {}
    total_mesures = 0
    total_alerts = 0
    async for m in db.mesures.find(
            {"chantier_id": {"$in": company_chantier_ids}},
            {"_id": 0, "chantier_id": 1, "alerts": 1}):
        total_mesures += 1
        alerts = len(m.get("alerts") or [])
        total_alerts += alerts
        tech = chantier_to_tech.get(m["chantier_id"]) or "unassigned"
        slot = by_tech.setdefault(tech, {"mesures": 0, "alerts": 0})
        slot["mesures"] += 1
        slot["alerts"] += alerts

    tech_users: dict[str, dict] = {}
    async for u in db.users.find(
            {"company_id": company}, {"_id": 0, "id": 1, "name": 1, "role": 1}):
        tech_users[u["id"]] = u

    tech_breakdown = []
    for tid, stats in by_tech.items():
        info = tech_users.get(tid)
        tech_breakdown.append({
            "user_id": tid,
            "name": info["name"] if info else "Non affecté",
            "role": info["role"] if info else "—",
            "mesures": stats["mesures"],
            "alerts": stats["alerts"],
        })
    tech_breakdown.sort(key=lambda x: x["mesures"], reverse=True)

    return {
        "total_chantiers": total,
        "by_status": by_status,
        "closure_rate": closure_rate,
        "total_mesures": total_mesures,
        "total_alerts": total_alerts,
        "by_technician": tech_breakdown,
    }


@api.get("/stats/commercials")
async def stats_commercials(user=Depends(require_admin)):
    company = user.get("company_id", "default")
    commercials = await db.users.find(
        {"company_id": company, "role": "commercial"},
        {"_id": 0, "id": 1, "name": 1, "email": 1}
    ).to_list(500)
    rows = []
    total_created = 0
    total_converted = 0
    for u in commercials:
        created = await db.chantiers.count_documents(
            {"company_id": company, "created_by": u["id"]})
        converted = await db.chantiers.count_documents(
            {"company_id": company, "created_by": u["id"],
             "status": {"$in": list(CONVERTED_STATUSES)}})
        rate = round((converted / created) * 100, 1) if created else 0.0
        rows.append({
            "user_id": u["id"],
            "name": u["name"],
            "email": u["email"],
            "created": created,
            "converted": converted,
            "conversion_rate": rate,
        })
        total_created += created
        total_converted += converted
    rows.sort(key=lambda r: r["conversion_rate"], reverse=True)
    global_rate = round((total_converted / total_created) * 100, 1) if total_created else 0.0
    return {
        "commercials": rows,
        "total_created": total_created,
        "total_converted": total_converted,
        "global_conversion_rate": global_rate,
    }


@api.get("/stats/commercials/export.pdf")
async def stats_commercials_pdf(user=Depends(require_admin)):
    data = await stats_commercials(user)  # reuse
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, title="Rapport Performance Commerciaux")
    styles = getSampleStyleSheet()
    story: list[Any] = []
    story.append(Paragraph("<b>MesureChâssis</b> — Rapport Performance Commerciaux",
                            styles["Title"]))
    story.append(Spacer(1, 14))
    story.append(Paragraph(
        f"<b>Société :</b> {user.get('company_id', 'default')}", styles["Normal"]))
    story.append(Paragraph(
        f"<b>Date :</b> {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}",
        styles["Normal"]))
    story.append(Spacer(1, 8))
    story.append(Paragraph(
        f"<b>Total chantiers créés :</b> {data['total_created']}  ·  "
        f"<b>Convertis :</b> {data['total_converted']}  ·  "
        f"<b>Taux global :</b> {data['global_conversion_rate']}%",
        styles["Normal"]))
    story.append(Spacer(1, 18))
    rows = [["Commercial", "Email", "Créés", "Convertis", "Conversion %"]]
    for r in data["commercials"]:
        rows.append([r["name"], r["email"], str(r["created"]),
                     str(r["converted"]), f"{r['conversion_rate']}%"])
    tbl = Table(rows, colWidths=[110, 170, 60, 70, 80])
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#FF5A00")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
    ]))
    story.append(tbl)
    doc.build(story)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/pdf",
        headers={"Content-Disposition":
                 'attachment; filename="rapport-performance-commerciaux.pdf"'})


# --- Export routes -------------------------------------------------------
def _status_label(s: str) -> str:
    return {"devis_a_faire": "Devis à faire",
            "technique_a_valider": "Technique à valider",
            "en_commande": "En commande",
            "en_fabrication": "En fabrication",
            "cloture": "Clôturé"}.get(s, s)


def _block_label(b: str) -> str:
    return {"standard": "Standard", "coulissant": "Coulissant",
            "porte": "Porte", "trapeze": "Trapèze"}.get(b, b)


@api.get("/chantiers/{chantier_id}/export.pdf")
async def export_pdf(chantier_id: str, user=Depends(auth_user)):
    chantier = await db.chantiers.find_one(
        {"id": chantier_id, "company_id": user.get("company_id", "default")},
        {"_id": 0})
    if not chantier:
        raise HTTPException(404, "Chantier introuvable")
    mesures = await db.mesures.find({"chantier_id": chantier_id}, {"_id": 0}) \
        .sort("created_at", 1).to_list(500)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, title=f"MesureChâssis - {chantier['client_name']}")
    styles = getSampleStyleSheet()
    story: list[Any] = []

    story.append(Paragraph("<b>MesureChâssis</b> — Fiche Chantier", styles["Title"]))
    story.append(Spacer(1, 12))
    story.append(Paragraph(f"<b>Client :</b> {chantier['client_name']}", styles["Normal"]))
    story.append(Paragraph(f"<b>Adresse :</b> {chantier['address']}", styles["Normal"]))
    story.append(Paragraph(f"<b>Statut :</b> {_status_label(chantier['status'])}", styles["Normal"]))
    story.append(Paragraph(f"<b>Date :</b> {chantier['created_at'][:10]}", styles["Normal"]))
    story.append(Spacer(1, 18))
    story.append(Paragraph(f"<b>Ouvertures ({len(mesures)})</b>", styles["Heading2"]))

    for i, m in enumerate(mesures, 1):
        story.append(Spacer(1, 8))
        story.append(Paragraph(
            f"<b>#{i} — {m.get('label', '')} ({_block_label(m['block_type'])})</b>",
            styles["Heading3"]))
        rows = [["Mesure", "Valeur (mm)"]]
        fields = [
            ("Hauteur baie", m.get("bay_height")),
            ("Largeur baie", m.get("bay_width")),
            ("Diagonale 1", m.get("bay_diagonal_1") or m.get("bay_diagonal")),
            ("Diagonale 2", m.get("bay_diagonal_2") or m.get("bay_diagonal")),
            ("Réserve Sol Fini", m.get("floor_reserve")),
            ("Épaisseur Bloc Béton", m.get("bloc_thickness")),
            ("Épaisseur Isolant", m.get("insulation_thickness")),
            ("Finition extérieure", m.get("finish_outer")),
            ("Finition intérieure", m.get("finish_inner")),
            ("Type paroi", m.get("wall_type")),
            # Legacy
            ("Largeur haut", m.get("width_top")),
            ("Largeur milieu", m.get("width_middle")),
            ("Largeur bas", m.get("width_bottom")),
            ("Hauteur gauche", m.get("height_left")),
            ("Hauteur milieu", m.get("height_middle")),
            ("Hauteur droite", m.get("height_right")),
            ("Diagonale 1", m.get("diag_1")),
            ("Diagonale 2", m.get("diag_2")),
            ("Hauteur 1/4 gauche", m.get("height_quarter_left")),
            ("Hauteur 1/4 droite", m.get("height_quarter_right")),
            ("Hauteur petite", m.get("height_small")),
            ("Hauteur grande", m.get("height_large")),
            ("Largeur petite", m.get("width_small")),
            ("Largeur intermédiaire", m.get("width_intermediate")),
        ]
        for label, val in fields:
            if val is not None:
                if label == "Type paroi":
                    label_map = {"ite": "ITE", "iti": "ITI",
                                "brique_parement": "Brique de parement",
                                "crepi_simple": "Crépi simple"}
                    rows.append([label, label_map.get(str(val), str(val))])
                else:
                    rows.append([label, str(val)])
        if m.get("slope_angle_deg") is not None:
            rows.append(["Angle de pente", f"{m['slope_angle_deg']}°"])
        tbl = Table(rows, colWidths=[260, 200])
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#FF5A00")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ]))
        story.append(tbl)
        if m.get("alerts"):
            story.append(Spacer(1, 4))
            for a in m["alerts"]:
                story.append(Paragraph(
                    f'<font color="#CC0000">{a}</font>', styles["Normal"]))

    doc.build(story)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition":
                                      f'attachment; filename="chantier-{chantier_id}.pdf"'})


@api.get("/chantiers/{chantier_id}/export.json")
async def export_json(chantier_id: str, user=Depends(auth_user)):
    chantier = await db.chantiers.find_one(
        {"id": chantier_id, "company_id": user.get("company_id", "default")},
        {"_id": 0})
    if not chantier:
        raise HTTPException(404, "Chantier introuvable")
    mesures = await db.mesures.find({"chantier_id": chantier_id}, {"_id": 0}) \
        .sort("created_at", 1).to_list(500)

    def _mesure_struct(m: dict) -> dict:
        bt = m.get("block_type")
        common = {
            "id": m.get("id"),
            "label": m.get("label"),
            "block_type": bt,
            "created_at": m.get("created_at"),
        }
        if bt == "trapeze":
            return {**common, "shape": "trapezoidal", "dimensions_mm": {
                "width": m.get("bay_width"),
                "height_left": m.get("height_left"),
                "height_right": m.get("height_right"),
            }}
        # rectangular family
        dims = {
            "width": m.get("bay_width"),
            "height": m.get("bay_height"),
            "diagonal_1": m.get("bay_diagonal_1") or m.get("bay_diagonal"),
            "diagonal_2": m.get("bay_diagonal_2") or m.get("bay_diagonal"),
        }
        if bt in ("porte", "coulissant"):
            dims["floor_reserve"] = m.get("floor_reserve")
        return {**common, "shape": "rectangular", "dimensions_mm": dims,
                "diagonals_verified": {
                    "d1": bool(m.get("diag_1_verified")),
                    "d2": bool(m.get("diag_2_verified")),
                }}

    return {
        "schema_version": "mc.v1",
        "exported_at": datetime.now(timezone.utc).isoformat(),
        "company_id": chantier.get("company_id"),
        "client": {
            "display_name": chantier.get("client_name"),
            "first_name": chantier.get("first_name"),
            "last_name": chantier.get("last_name"),
            "address": chantier.get("address"),
            "postal_code": chantier.get("postal_code"),
            "city": chantier.get("city"),
        },
        "project": {
            "id": chantier.get("id"),
            "status": chantier.get("status"),
            "appointment_at": chantier.get("appointment_at"),
            "notes": chantier.get("notes"),
            "created_at": chantier.get("created_at"),
            "assigned_to": chantier.get("assigned_to"),
        },
        "openings_count": len(mesures),
        "openings": [_mesure_struct(m) for m in mesures],
    }


@api.get("/chantiers/{chantier_id}/export.csv")
async def export_csv(chantier_id: str, user=Depends(auth_user)):
    """Plain tabular CSV — manufacturing / cutting machinery friendly."""
    import csv
    import io
    chantier = await db.chantiers.find_one(
        {"id": chantier_id, "company_id": user.get("company_id", "default")},
        {"_id": 0})
    if not chantier:
        raise HTTPException(404, "Chantier introuvable")
    mesures = await db.mesures.find({"chantier_id": chantier_id}, {"_id": 0}) \
        .sort("created_at", 1).to_list(500)

    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=";", quoting=csv.QUOTE_MINIMAL)
    # Header
    writer.writerow([
        "Chantier", "Adresse", "Code Postal", "Ville", "Statut",
        "Label", "Type", "Forme",
        "Largeur (mm)", "Hauteur (mm)", "Hauteur G (mm)", "Hauteur D (mm)",
        "Diag 1 (mm)", "Diag 2 (mm)", "Diag1 OK", "Diag2 OK",
        "Réserve sol (mm)", "Épaisseur bloc (mm)", "Paroi",
        "Date mesure",
    ])
    client_disp = chantier.get("client_name") or "—"
    for m in mesures:
        bt = m.get("block_type") or "—"
        is_trap = bt == "trapeze"
        writer.writerow([
            client_disp,
            chantier.get("address") or "",
            chantier.get("postal_code") or "",
            chantier.get("city") or "",
            chantier.get("status") or "",
            m.get("label") or "",
            bt,
            "trapezoidal" if is_trap else "rectangular",
            m.get("bay_width") or "",
            "" if is_trap else (m.get("bay_height") or ""),
            m.get("height_left") or "" if is_trap else "",
            m.get("height_right") or "" if is_trap else "",
            "" if is_trap else (m.get("bay_diagonal_1") or m.get("bay_diagonal") or ""),
            "" if is_trap else (m.get("bay_diagonal_2") or m.get("bay_diagonal") or ""),
            "" if is_trap else ("oui" if m.get("diag_1_verified") else "non"),
            "" if is_trap else ("oui" if m.get("diag_2_verified") else "non"),
            m.get("floor_reserve") or "",
            m.get("bloc_thickness") or "",
            m.get("wall_type") or "",
            (m.get("created_at") or "")[:19].replace("T", " "),
        ])
    content = buf.getvalue().encode("utf-8-sig")  # BOM for Excel-FR compatibility
    safe = (chantier.get("client_name") or chantier_id).replace(" ", "_").replace("/", "-")
    return Response(
        content=content,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="MesureChassis_{safe}.csv"'},
    )


@api.get("/chantiers/{chantier_id}/export.xlsx")
async def export_xlsx(chantier_id: str, user=Depends(auth_user)):
    chantier = await db.chantiers.find_one(
        {"id": chantier_id, "company_id": user.get("company_id", "default")},
        {"_id": 0})
    if not chantier:
        raise HTTPException(404, "Chantier introuvable")
    mesures = await db.mesures.find({"chantier_id": chantier_id}, {"_id": 0}) \
        .sort("created_at", 1).to_list(500)

    wb = Workbook()
    info = wb.active
    info.title = "Chantier"
    head = Font(bold=True, color="FFFFFF")
    fill = PatternFill(start_color="FF5A00", end_color="FF5A00", fill_type="solid")
    info["A1"] = "MesureChâssis — Fiche Chantier"
    info["A1"].font = Font(bold=True, size=14)
    pairs = [
        ("Client", chantier["client_name"]),
        ("Adresse", chantier["address"]),
        ("Statut", _status_label(chantier["status"])),
        ("Date", chantier["created_at"][:10]),
        ("Signé le", chantier.get("signed_at") or "—"),
    ]
    for i, (k, v) in enumerate(pairs, start=3):
        info.cell(row=i, column=1, value=k).font = Font(bold=True)
        info.cell(row=i, column=2, value=v)
    info.column_dimensions["A"].width = 22
    info.column_dimensions["B"].width = 50

    ws = wb.create_sheet("Mesures")
    columns = [
        ("Libellé", "label"),
        ("Type bloc", "block_type"),
        ("H baie (mm)", "bay_height"),
        ("L baie (mm)", "bay_width"),
        ("Diag (mm)", "bay_diagonal"),
        ("Réserve sol fini (mm)", "floor_reserve"),
        ("Épais. bloc béton (mm)", "bloc_thickness"),
        ("Type paroi", "wall_type"),
        ("Épais. isolant (mm)", "insulation_thickness"),
        ("Finition ext. (mm)", "finish_outer"),
        ("Finition int. (mm)", "finish_inner"),
        ("Angle pente (°)", "slope_angle_deg"),
        ("Alertes", "alerts"),
    ]
    for col_idx, (label, _) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = head
        cell.fill = fill
    wall_map = {"ite": "ITE", "iti": "ITI",
                "brique_parement": "Brique de parement",
                "crepi_simple": "Crépi simple"}
    block_map = {"standard": "Standard", "coulissant": "Coulissant",
                 "porte": "Porte", "trapeze": "Trapèze"}
    for row_idx, m in enumerate(mesures, start=2):
        for col_idx, (_, key) in enumerate(columns, start=1):
            v: Any = m.get(key)
            if key == "wall_type" and v:
                v = wall_map.get(v, v)
            elif key == "block_type" and v:
                v = block_map.get(v, v)
            elif key == "alerts":
                v = " ; ".join(v) if v else ""
            ws.cell(row=row_idx, column=col_idx, value=v)
    for c in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value is not None else 0) for cell in c)
        ws.column_dimensions[c[0].column_letter].width = min(max_len + 2, 36)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="chantier-{chantier_id}.xlsx"'},
    )


@api.post("/chantiers/{chantier_id}/signature", response_model=Chantier)
async def save_signature(chantier_id: str, payload: SignatureIn, user=Depends(auth_user)):
    company = user.get("company_id", "default")
    if not payload.signature.strip():
        raise HTTPException(400, "Signature vide")
    res = await db.chantiers.update_one(
        {"id": chantier_id, "company_id": company},
        {"$set": {
            "client_signature": payload.signature,
            "signed_at": datetime.now(timezone.utc).isoformat(),
        }},
    )
    if res.matched_count == 0:
        raise HTTPException(404, "Chantier introuvable")
    doc = await db.chantiers.find_one({"id": chantier_id}, {"_id": 0})
    return Chantier(**doc)


@api.delete("/chantiers/{chantier_id}/signature", response_model=Chantier)
async def delete_signature(chantier_id: str, user=Depends(auth_user)):
    company = user.get("company_id", "default")
    await db.chantiers.update_one(
        {"id": chantier_id, "company_id": company},
        {"$set": {"client_signature": None, "signed_at": None}},
    )
    doc = await db.chantiers.find_one({"id": chantier_id, "company_id": company}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Chantier introuvable")
    return Chantier(**doc)


# --- Demo data seeding ---------------------------------------------------
DEMO_USERS = [
    {"name": "Marc Dubois", "email": "admin@mesurechassis.fr",
     "password": "admin123", "role": "admin"},
    {"name": "Sophie Martin", "email": "commercial@mesurechassis.fr",
     "password": "commercial123", "role": "commercial"},
    {"name": "Lucas Petit", "email": "tech@mesurechassis.fr",
     "password": "tech123", "role": "technician"},
]


@app.on_event("startup")
async def seed_data():
    # Seed users (idempotent)
    user_ids: dict[str, str] = {}
    for u in DEMO_USERS:
        existing = await db.users.find_one({"email": u["email"]})
        if existing:
            user_ids[u["role"]] = existing["id"]
            continue
        uid = str(uuid.uuid4())
        await db.users.insert_one({
            "id": uid, "name": u["name"], "email": u["email"], "role": u["role"],
            "company_id": "default",
            "hashed_password": hash_password(u["password"]),
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
        user_ids[u["role"]] = uid
        logger.info("Seeded user %s", u["email"])

    # Seed chantiers if none exist
    if await db.chantiers.count_documents({}) == 0:
        demos = [
            ("Famille Lefèvre", "12 rue de la Paix, 75002 Paris", "devis_a_faire"),
            ("Boulangerie Moreau", "45 av. Victor Hugo, 69006 Lyon", "devis_a_faire"),
            ("M. et Mme Bernard", "8 chemin des Vignes, 33000 Bordeaux", "technique_a_valider"),
            ("SCI Le Clos", "23 rue Nationale, 59000 Lille", "technique_a_valider"),
            ("Cabinet Dr. Rousseau", "5 place Bellecour, 69002 Lyon", "cloture"),
        ]
        for name, addr, status_v in demos:
            await db.chantiers.insert_one({
                "id": str(uuid.uuid4()),
                "client_name": name,
                "address": addr,
                "status": status_v,
                "created_by": user_ids.get("commercial", "system"),
                "assigned_to": user_ids.get("technician"),
                "company_id": "default",
                "created_at": datetime.now(timezone.utc).isoformat(),
            })
        logger.info("Seeded %d demo chantiers", len(demos))

    # Backfill missing company_id on existing rows (from previous schema)
    await db.chantiers.update_many(
        {"company_id": {"$exists": False}}, {"$set": {"company_id": "default"}})
    await db.feedbacks.update_many(
        {"company_id": {"$exists": False}}, {"$set": {"company_id": "default"}})


# --- App wiring ----------------------------------------------------------
app.include_router(api)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
