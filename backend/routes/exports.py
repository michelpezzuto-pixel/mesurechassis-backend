"""Routes d'export d'un chantier : PDF (avec photos), CSV, XLSX, JSON, ERP générique."""
from __future__ import annotations

import base64
import csv
import io
from datetime import datetime, timezone
from typing import Any
from xml.sax.saxutils import escape as xml_escape

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import Response, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (Image as RLImage, Paragraph,
                                SimpleDocTemplate, Spacer, Table, TableStyle)

from db import db
from deps import require_active_subscription
from utils import WALL_TYPE_LABELS, block_label, status_label
from i18n_labels import pdf_labels, status_label_i18n

router = APIRouter()


def _safe_filename(name: str) -> str:
    """Latin-1 safe filename (HTTP headers ne supportent pas l'UTF-8 brut).

    Translitération basique des caractères français + suppression du reste.
    """
    import re
    import unicodedata
    n = unicodedata.normalize("NFKD", name)
    n = "".join(c for c in n if not unicodedata.combining(c))
    n = re.sub(r"[^\w\-_.]", "_", n)
    return n[:80] or "chantier"


def block_free_plan_exports(
    user=Depends(require_active_subscription),
) -> dict:
    """Bloque tous les exports pour le plan Free (anti-fraud).

    Le mode Artisan Unique ne bypass PAS ce verrou : un compte free reste free.
    Seul le passage en Pro débloque les exports.
    """
    if (user.get("plan") or "trial") == "free":
        raise HTTPException(
            status_code=402,
            detail={
                "code": "free_plan_no_export",
                "message": (
                    "Les exports (PDF, Excel, CSV, JSON) sont réservés "
                    "aux abonnés Pro. Passez en Pro pour débloquer les "
                    "exports techniques."
                ),
                "plan": "free",
            },
        )
    return user


def restrict_advanced_exports(
    user=Depends(block_free_plan_exports),
) -> dict:
    """Excel / CSV / JSON réservés au Technicien et Admin.

    Commercial sans Mode Artisan : 403 (uniquement PDF accessible).
    Le verrou Free passe d'abord (anti-fraud).
    """
    if user.get("artisan_mode"):
        return user
    if user["role"] == "commercial":
        raise HTTPException(
            status_code=403,
            detail=(
                "Les exports Excel/CSV/JSON sont réservés aux techniciens "
                "et administrateurs. Seul le PDF est disponible pour les "
                "commerciaux."
            ),
        )
    return user


# ---------------------------- PDF --------------------------------------
@router.get("/chantiers/{chantier_id}/export.pdf")
async def export_pdf(
    chantier_id: str,
    lang: str = "fr",
    user=Depends(block_free_plan_exports),
):
    """Export PDF avec support multilingue FR/NL/EN via `?lang=`.

    La langue détermine tous les libellés visibles (en-tête, statuts, champs
    de mesures, sections photos). Fallback FR si langue non reconnue.
    """
    L = pdf_labels(lang)
    chantier = await db.chantiers.find_one(
        {
            "id": chantier_id,
            "company_id": user.get("company_id", "default"),
        },
        {"_id": 0},
    )
    if not chantier:
        raise HTTPException(404, "Chantier introuvable")
    mesures = (
        await db.mesures.find({"chantier_id": chantier_id}, {"_id": 0})
        .sort("created_at", 1)
        .to_list(500)
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4, title=f"MesureChâssis - {chantier['client_name']}"
    )
    styles = getSampleStyleSheet()
    story: list[Any] = []

    # ---- En-tête : logo entreprise + branding + mention DOCUMENT INTERNE
    company_doc = await db.companies.find_one(
        {"company_id": user.get("company_id", "default")},
        {"_id": 0, "name": 1, "logo_base64": 1},
    ) or {}
    company_name = company_doc.get("name") or "MesureChâssis"
    logo_uri = company_doc.get("logo_base64") or ""

    header_logo_cell: Any = ""
    if logo_uri:
        try:
            if logo_uri.startswith("data:"):
                _, b64 = logo_uri.split(",", 1)
            else:
                b64 = logo_uri
            logo_bytes = base64.b64decode(b64)
            # Pré-validation PNG/JPG via PIL pour éviter qu'un fichier corrompu
            # ne crash doc.build() (lève "broken data stream").
            try:
                from PIL import Image as PILImage

                with PILImage.open(io.BytesIO(logo_bytes)) as _img:
                    _img.verify()
            except Exception:
                raise ValueError("logo image invalid")
            logo_io = io.BytesIO(logo_bytes)
            header_logo_cell = RLImage(
                logo_io, width=35 * mm, height=20 * mm, kind="proportional"
            )
        except Exception:
            header_logo_cell = ""

    header_right = [
        Paragraph(
            f'<font size="14"><b>{company_name}</b></font>',
            styles["Normal"],
        ),
        Spacer(1, 2),
        Paragraph(
            f'<font size="8" color="#666666">{L["via"]}</font>',
            styles["Normal"],
        ),
    ]
    header_tbl = Table(
        [[header_logo_cell, header_right]],
        colWidths=[40 * mm, 120 * mm],
    )
    header_tbl.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (0, 0), (0, 0), "LEFT"),
                ("ALIGN", (1, 0), (1, 0), "RIGHT"),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )
    story.append(header_tbl)
    story.append(Spacer(1, 6))
    # Bandeau orange "DOCUMENT INTERNE"
    internal_banner = Table(
        [[Paragraph(
            f'<font color="#FFFFFF" size="9">{L["internal_banner"]}</font>',
            styles["Normal"],
        )]],
        colWidths=[160 * mm],
    )
    internal_banner.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#FF5A00")),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    story.append(internal_banner)
    story.append(Spacer(1, 10))

    story.append(
        Paragraph(
            f"<b>{L['fiche']}</b>",
            styles["Title"],
        )
    )
    story.append(Spacer(1, 8))
    story.append(
        Paragraph(f"<b>{L['client']} :</b> {chantier['client_name']}", styles["Normal"])
    )
    story.append(
        Paragraph(f"<b>{L['address']} :</b> {chantier['address']}", styles["Normal"])
    )
    story.append(
        Paragraph(
            f"<b>{L['status']} :</b> {status_label_i18n(chantier['status'], lang)}",
            styles["Normal"],
        )
    )
    story.append(
        Paragraph(
            f"<b>{L['date']} :</b> {chantier['created_at'][:10]}", styles["Normal"]
        )
    )
    story.append(Spacer(1, 18))
    story.append(
        Paragraph(f"<b>{L['openings_section']} ({len(mesures)})</b>", styles["Heading2"])
    )

    for i, m in enumerate(mesures, 1):
        story.append(Spacer(1, 8))
        story.append(
            Paragraph(
                f"<b>#{i} — {m.get('label', '')} ({block_label(m['block_type'], m.get('options'))})</b>",
                styles["Heading3"],
            )
        )
        rows = [[L["measure_col"], L["value_col"]]]
        opts = m.get("options") or {}
        fields = [
            (L["field"]["bay_height"], m.get("bay_height")),
            (L["field"]["bay_width"], m.get("bay_width")),
            (L["field"]["bay_diagonal_1"], m.get("bay_diagonal_1") or m.get("bay_diagonal")),
            (L["field"]["bay_diagonal_2"], m.get("bay_diagonal_2") or m.get("bay_diagonal")),
            (L["field"]["floor_reserve"], m.get("floor_reserve")),
            (L["field"]["bloc_thickness"], m.get("bloc_thickness")),
            (L["field"]["insulation_thickness"], m.get("insulation_thickness")),
            (L["field"]["finish_outer"], m.get("finish_outer")),
            (L["field"]["finish_inner"], m.get("finish_inner")),
            (L["field"]["wall_type"], m.get("wall_type")),
            # === Champs options (Étape 3 — wizard V2) ===
            (L["field"]["breastwork_present"], L["yes"] if opts.get("has_breastwork") else None),
            (L["field"]["breastwork_height"], opts.get("breastwork_height_mm")),
            (L["field"]["feuillure_left"], opts.get("feuillure_left_mm")),
            (L["field"]["feuillure_right"], opts.get("feuillure_right_mm")),
            (L["field"]["feuillure_top"], opts.get("feuillure_top_mm")),
            (L["field"]["sill_installed"], L["yes"] if opts.get("sill_already_installed") is True else (L["no"] if opts.get("sill_already_installed") is False else None)),
            (L["field"]["sill_thickness"], opts.get("sill_thickness_mm")),
            (L["field"]["horizontal_cut"], L["yes"] if opts.get("has_horizontal_cut") else None),
            (L["field"]["mark_1m_active"], L["yes"] if opts.get("has_1m_level_mark") else None),
            (L["field"]["mark_1m_brut"], opts.get("trait_1m_brut_mm")),
            # Spécifiques formes
            (L["field"]["triangle_base"], opts.get("triangle_base_mm")),
            (L["field"]["triangle_height"], opts.get("triangle_height_mm")),
            (L["field"]["oeil_diameter"], opts.get("oeil_diameter_mm")),
            (L["field"]["garage_lintel"], opts.get("garage_lintel_mm")),
            (L["field"]["garage_ecoincon_left"], opts.get("garage_ecoincon_left_mm")),
            (L["field"]["garage_ecoincon_right"], opts.get("garage_ecoincon_right_mm")),
            # Legacy
            (L["field"]["width_top"], m.get("width_top")),
            (L["field"]["width_middle"], m.get("width_middle")),
            (L["field"]["width_bottom"], m.get("width_bottom")),
            (L["field"]["height_left"], m.get("height_left")),
            (L["field"]["height_middle"], m.get("height_middle")),
            (L["field"]["height_right"], m.get("height_right")),
            (L["field"]["diag_1"], m.get("diag_1")),
            (L["field"]["diag_2"], m.get("diag_2")),
            (L["field"]["height_quarter_left"], m.get("height_quarter_left")),
            (L["field"]["height_quarter_right"], m.get("height_quarter_right")),
            (L["field"]["height_small"], m.get("height_small")),
            (L["field"]["height_large"], m.get("height_large")),
            (L["field"]["width_small"], m.get("width_small")),
            (L["field"]["width_intermediate"], m.get("width_intermediate")),
        ]
        for label, val in fields:
            if val is not None:
                if label == L["field"]["wall_type"]:
                    rows.append(
                        [label, WALL_TYPE_LABELS.get(str(val), str(val))]
                    )
                else:
                    rows.append([label, str(val)])
        if m.get("slope_angle_deg") is not None:
            rows.append([L["slope_angle"], f"{m['slope_angle_deg']}°"])
        tbl = Table(rows, colWidths=[260, 200])
        tbl.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#FF5A00")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ]
            )
        )
        story.append(tbl)
        if m.get("alerts"):
            story.append(Spacer(1, 4))
            for a in m["alerts"]:
                story.append(
                    Paragraph(
                        f'<font color="#CC0000">{a}</font>', styles["Normal"]
                    )
                )

    # --- Photos site (Anti-litige) ------------------------------------
    site_photos = chantier.get("site_photos") or []
    if site_photos:
        story.append(Spacer(1, 18))
        story.append(
            Paragraph(
                f"<b>{L['photos_section']} ({len(site_photos)})</b>",
                styles["Heading2"],
            )
        )
        story.append(
            Paragraph(
                f'<font size="9" color="#666666"><i>{L["photos_caption"]}</i></font>',
                styles["Normal"],
            )
        )
        story.append(Spacer(1, 8))

        caption_style = ParagraphStyle(
            "PhotoCaption",
            parent=styles["Normal"],
            fontName="Helvetica-Oblique",
            fontSize=9,
            textColor=colors.HexColor("#444444"),
            alignment=1,
            leading=11,
        )
        photo_cells: list[list[Any]] = []
        row: list[Any] = []
        for idx, ph in enumerate(site_photos, 1):
            uri = ph.get("uri") or ""
            caption = (ph.get("caption") or "").strip() or L["photo_default_caption"].format(idx=idx)
            cell_content: list[Any] = []
            try:
                if uri.startswith("data:"):
                    _, b64 = uri.split(",", 1)
                else:
                    b64 = uri
                img_bytes = base64.b64decode(b64)
                img_io = io.BytesIO(img_bytes)
                img = RLImage(
                    img_io, width=80 * mm, height=60 * mm, kind="proportional"
                )
                cell_content.append(img)
            except Exception:
                cell_content.append(
                    Paragraph(
                        f'<font color="#999999">[{L["photo_unreadable"].format(idx=idx)}]</font>',
                        styles["Normal"],
                    )
                )
            cell_content.append(Spacer(1, 4))
            cell_content.append(
                Paragraph(f"#{idx} — {caption}", caption_style)
            )
            row.append(cell_content)
            if len(row) == 2:
                photo_cells.append(row)
                row = []
        if row:
            row.append("")
            photo_cells.append(row)
        if photo_cells:
            photo_table = Table(photo_cells, colWidths=[90 * mm, 90 * mm])
            photo_table.setStyle(
                TableStyle(
                    [
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ("LEFTPADDING", (0, 0), (-1, -1), 4),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                        ("TOPPADDING", (0, 0), (-1, -1), 6),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
                    ]
                )
            )
            story.append(photo_table)

    doc.build(story)
    buf.seek(0)
    safe = _safe_filename(chantier.get("client_name") or chantier_id)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={
            "Content-Disposition": (
                f'attachment; filename="MesureChassis_{safe}.pdf"'
            )
        },
    )


# ---------------------------- JSON -------------------------------------
@router.get("/chantiers/{chantier_id}/export.json")
async def export_json(
    chantier_id: str, user=Depends(restrict_advanced_exports)
):
    chantier = await db.chantiers.find_one(
        {
            "id": chantier_id,
            "company_id": user.get("company_id", "default"),
        },
        {"_id": 0},
    )
    if not chantier:
        raise HTTPException(404, "Chantier introuvable")
    mesures = (
        await db.mesures.find({"chantier_id": chantier_id}, {"_id": 0})
        .sort("created_at", 1)
        .to_list(500)
    )

    def _opening(m: dict) -> dict:
        bt = m.get("block_type")
        common = {
            "id": m.get("id"),
            "label": m.get("label"),
            "block_type": bt,
            "block_label": block_label(bt or "", m.get("options")),
            "created_at": m.get("created_at"),
            "alerts": m.get("alerts") or [],
            "slope_angle_deg": m.get("slope_angle_deg"),
            "renovation_mode": bool(m.get("renovation_mode")),
            "options": m.get("options") or {},
            "photo_url": m.get("photo_url"),
        }
        if bt == "trapeze":
            return {
                **common,
                "shape": "trapezoidal",
                "dimensions_mm": {
                    "width": m.get("bay_width"),
                    "height_left": m.get("height_left"),
                    "height_right": m.get("height_right"),
                    # Données complémentaires si saisies (legacy trapèze)
                    "width_small": m.get("width_small"),
                    "width_intermediate": m.get("width_intermediate"),
                    "height_small": m.get("height_small"),
                    "height_large": m.get("height_large"),
                },
                "construction": _construction(m),
            }
        dims = {
            "width": m.get("bay_width"),
            "height": m.get("bay_height"),
            "diagonal_1": m.get("bay_diagonal_1") or m.get("bay_diagonal"),
            "diagonal_2": m.get("bay_diagonal_2") or m.get("bay_diagonal"),
        }
        if bt in ("porte", "coulissant"):
            dims["floor_reserve"] = m.get("floor_reserve")
        # Renovation mode : 4 cotes explicites
        if m.get("renovation_mode"):
            dims["width_top"] = m.get("width_top")
            dims["width_bottom"] = m.get("width_bottom")
            dims["height_left"] = m.get("height_left")
            dims["height_right"] = m.get("height_right")
        # Legacy fields (gardés si non-null pour rétrocompat)
        legacy = {
            k: m.get(k)
            for k in (
                "width_top", "width_middle", "width_bottom",
                "height_left", "height_middle", "height_right",
                "diag_1", "diag_2",
                "height_quarter_left", "height_quarter_right",
            )
            if m.get(k) is not None and k not in dims
        }
        if legacy:
            dims["legacy"] = legacy
        return {
            **common,
            "shape": "rectangular",
            "dimensions_mm": dims,
            "diagonals_verified": {
                "d1": bool(m.get("diag_1_verified")),
                "d2": bool(m.get("diag_2_verified")),
            },
            "construction": _construction(m),
        }

    site_photos = chantier.get("site_photos") or []
    return {
        "schema_version": "mc.v2",
        "exported_at": datetime.now(timezone.utc).isoformat(),
        "company_id": chantier.get("company_id"),
        "client": {
            "display_name": chantier.get("client_name"),
            "first_name": chantier.get("first_name"),
            "last_name": chantier.get("last_name"),
            "address": chantier.get("address"),
            "postal_code": chantier.get("postal_code"),
            "city": chantier.get("city"),
        },
        "project": {
            "id": chantier.get("id"),
            "status": chantier.get("status"),
            "status_label": status_label(chantier.get("status") or ""),
            "appointment_at": chantier.get("appointment_at"),
            "notes": chantier.get("notes"),
            "created_at": chantier.get("created_at"),
            "assigned_to": chantier.get("assigned_to"),
        },
        "openings_count": len(mesures),
        "openings": [_opening(m) for m in mesures],
        "site_photos_count": len(site_photos),
        "site_photos": [
            {
                "index": i + 1,
                "caption": (p.get("caption") or "").strip(),
                "uri": p.get("uri"),
            }
            for i, p in enumerate(site_photos)
        ],
    }


def _construction(m: dict) -> dict:
    """Bloc maçonnerie + isolation (commun à toutes les formes)."""
    return {
        "bloc_thickness_mm": m.get("bloc_thickness"),
        "wall_type": m.get("wall_type"),
        "wall_type_label": (
            WALL_TYPE_LABELS.get(m.get("wall_type") or "", "")
        ),
        "insulation_thickness_mm": m.get("insulation_thickness"),
        "finish_outer_mm": m.get("finish_outer"),
        "finish_inner_mm": m.get("finish_inner"),
    }


# ---------------------------- CSV --------------------------------------
@router.get("/chantiers/{chantier_id}/export.csv")
async def export_csv(
    chantier_id: str, user=Depends(restrict_advanced_exports)
):
    """CSV tabulaire complet — atelier / machines de découpe."""
    chantier = await db.chantiers.find_one(
        {
            "id": chantier_id,
            "company_id": user.get("company_id", "default"),
        },
        {"_id": 0},
    )
    if not chantier:
        raise HTTPException(404, "Chantier introuvable")
    mesures = (
        await db.mesures.find({"chantier_id": chantier_id}, {"_id": 0})
        .sort("created_at", 1)
        .to_list(500)
    )

    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=";", quoting=csv.QUOTE_MINIMAL)
    # En-tête complet
    writer.writerow(
        [
            "Chantier", "Adresse", "Code Postal", "Ville", "Statut",
            "Label", "Type", "Forme", "Rénovation",
            # Dimensions principales
            "Largeur baie (mm)", "Hauteur baie (mm)",
            "Diag 1 (mm)", "Diag 2 (mm)", "Diag1 OK", "Diag2 OK",
            # Rénovation 4 cotes
            "L. haut (mm)", "L. bas (mm)",
            "H. gauche (mm)", "H. droite (mm)",
            # Legacy multi-points
            "L. milieu (mm)", "H. milieu (mm)",
            # Trapèze
            "L. petite (mm)", "L. inter (mm)",
            "H. petite (mm)", "H. grande (mm)",
            # Construction
            "Réserve sol (mm)", "Épaisseur bloc (mm)",
            "Paroi", "Épais. isolant (mm)",
            "Finition ext (mm)", "Finition int (mm)",
            "Angle pente (°)",
            "Alertes",
            "Date mesure",
        ]
    )
    client_disp = chantier.get("client_name") or "—"
    for m in mesures:
        bt = m.get("block_type") or "—"
        is_trap = bt == "trapeze"
        is_renov = bool(m.get("renovation_mode"))
        writer.writerow(
            [
                client_disp,
                chantier.get("address") or "",
                chantier.get("postal_code") or "",
                chantier.get("city") or "",
                status_label(chantier.get("status") or ""),
                m.get("label") or "",
                block_label(bt, m.get("options")),
                "trapezoidal" if is_trap else "rectangular",
                "oui" if is_renov else "non",
                m.get("bay_width") or "",
                "" if is_trap else (m.get("bay_height") or ""),
                "" if is_trap else (
                    m.get("bay_diagonal_1") or m.get("bay_diagonal") or ""
                ),
                "" if is_trap else (
                    m.get("bay_diagonal_2") or m.get("bay_diagonal") or ""
                ),
                "" if is_trap else (
                    "oui" if m.get("diag_1_verified") else "non"
                ),
                "" if is_trap else (
                    "oui" if m.get("diag_2_verified") else "non"
                ),
                m.get("width_top") or "",
                m.get("width_bottom") or "",
                m.get("height_left") or "",
                m.get("height_right") or "",
                m.get("width_middle") or "",
                m.get("height_middle") or "",
                m.get("width_small") or "",
                m.get("width_intermediate") or "",
                m.get("height_small") or "",
                m.get("height_large") or "",
                m.get("floor_reserve") or "",
                m.get("bloc_thickness") or "",
                WALL_TYPE_LABELS.get(m.get("wall_type") or "", "") or "",
                m.get("insulation_thickness") or "",
                m.get("finish_outer") or "",
                m.get("finish_inner") or "",
                m.get("slope_angle_deg") or "",
                " ; ".join(m.get("alerts") or []),
                (m.get("created_at") or "")[:19].replace("T", " "),
            ]
        )
    # --- Bloc photos en fin de fichier (anti-litige) ---------------------
    site_photos = chantier.get("site_photos") or []
    if site_photos:
        writer.writerow([])
        writer.writerow(["[PHOTOS ANTI-LITIGE]"])
        writer.writerow(["#", "Légende", "Format"])
        for i, ph in enumerate(site_photos, 1):
            cap = (ph.get("caption") or "").strip() or f"Photo {i}"
            uri = ph.get("uri") or ""
            fmt = "base64 data URI" if uri.startswith("data:") else "URL"
            writer.writerow([i, cap, fmt])

    content = buf.getvalue().encode("utf-8-sig")
    safe = _safe_filename(chantier.get("client_name") or chantier_id)
    return Response(
        content=content,
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": (
                f'attachment; filename="MesureChassis_{safe}.csv"'
            )
        },
    )


# ---------------------------- XLSX -------------------------------------
@router.get("/chantiers/{chantier_id}/export.xlsx")
async def export_xlsx(
    chantier_id: str, user=Depends(restrict_advanced_exports)
):
    chantier = await db.chantiers.find_one(
        {
            "id": chantier_id,
            "company_id": user.get("company_id", "default"),
        },
        {"_id": 0},
    )
    if not chantier:
        raise HTTPException(404, "Chantier introuvable")
    mesures = (
        await db.mesures.find({"chantier_id": chantier_id}, {"_id": 0})
        .sort("created_at", 1)
        .to_list(500)
    )

    wb = Workbook()
    info = wb.active
    info.title = "Chantier"
    head = Font(bold=True, color="FFFFFF")
    fill = PatternFill(
        start_color="FF5A00", end_color="FF5A00", fill_type="solid"
    )
    info["A1"] = "MesureChâssis — Fiche Chantier"
    info["A1"].font = Font(bold=True, size=14)
    pairs = [
        ("Client", chantier["client_name"]),
        ("Adresse", chantier["address"]),
        ("Code postal", chantier.get("postal_code") or "—"),
        ("Ville", chantier.get("city") or "—"),
        ("Statut", status_label(chantier["status"])),
        ("Date création", chantier["created_at"][:10]),
        ("Rendez-vous", chantier.get("appointment_at") or "—"),
        ("Notes", chantier.get("notes") or "—"),
        ("Nb. ouvertures", len(mesures)),
        ("Nb. photos site", len(chantier.get("site_photos") or [])),
    ]
    for i, (k, v) in enumerate(pairs, start=3):
        info.cell(row=i, column=1, value=k).font = Font(bold=True)
        info.cell(row=i, column=2, value=v)
    info.column_dimensions["A"].width = 24
    info.column_dimensions["B"].width = 60

    # --- Feuille Mesures (toutes les cotes) ------------------------------
    ws = wb.create_sheet("Mesures")
    columns = [
        ("Libellé", "label"),
        ("Type bloc", "block_type"),
        ("Forme", "_shape"),
        ("Rénovation", "_renov"),
        # Baie brute
        ("L baie (mm)", "bay_width"),
        ("H baie (mm)", "bay_height"),
        ("Diag 1 (mm)", "bay_diagonal_1"),
        ("Diag 2 (mm)", "bay_diagonal_2"),
        ("Diag1 OK", "diag_1_verified"),
        ("Diag2 OK", "diag_2_verified"),
        # Rénovation 4 cotes
        ("L. haut (mm)", "width_top"),
        ("L. bas (mm)", "width_bottom"),
        ("H. gauche (mm)", "height_left"),
        ("H. droite (mm)", "height_right"),
        # Multi-points (coulissant / legacy)
        ("L. milieu (mm)", "width_middle"),
        ("H. milieu (mm)", "height_middle"),
        ("H. 1/4 G (mm)", "height_quarter_left"),
        ("H. 1/4 D (mm)", "height_quarter_right"),
        # Trapèze
        ("L. petite (mm)", "width_small"),
        ("L. inter (mm)", "width_intermediate"),
        ("H. petite (mm)", "height_small"),
        ("H. grande (mm)", "height_large"),
        # Construction
        ("Réserve sol (mm)", "floor_reserve"),
        ("Épais. bloc béton (mm)", "bloc_thickness"),
        ("Type paroi", "wall_type"),
        ("Épais. isolant (mm)", "insulation_thickness"),
        ("Finition ext (mm)", "finish_outer"),
        ("Finition int (mm)", "finish_inner"),
        ("Angle pente (°)", "slope_angle_deg"),
        ("Alertes", "alerts"),
        ("Date mesure", "created_at"),
    ]
    for col_idx, (label, _) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = head
        cell.fill = fill
    for row_idx, m in enumerate(mesures, start=2):
        for col_idx, (_, key) in enumerate(columns, start=1):
            if key == "_shape":
                v: Any = "trapezoidal" if m.get("block_type") == "trapeze" else "rectangular"
            elif key == "_renov":
                v = "oui" if m.get("renovation_mode") else "non"
            else:
                v = m.get(key)
            if key == "wall_type" and v:
                v = WALL_TYPE_LABELS.get(v, v)
            elif key == "block_type" and v:
                v = block_label(v, m.get("options"))
            elif key == "alerts":
                v = " ; ".join(v) if v else ""
            elif key in ("diag_1_verified", "diag_2_verified"):
                v = "oui" if v else "non"
            elif key == "created_at" and v:
                v = str(v)[:19].replace("T", " ")
            ws.cell(row=row_idx, column=col_idx, value=v)
    for c in ws.columns:
        max_len = max(
            (len(str(cell.value)) if cell.value is not None else 0)
            for cell in c
        )
        ws.column_dimensions[c[0].column_letter].width = min(max_len + 2, 36)

    # --- Feuille Photos site (anti-litige) -------------------------------
    site_photos = chantier.get("site_photos") or []
    if site_photos:
        ph_ws = wb.create_sheet("Photos site")
        for col_idx, label in enumerate(
            ["#", "Légende", "Format", "Taille (caractères)"], start=1
        ):
            cell = ph_ws.cell(row=1, column=col_idx, value=label)
            cell.font = head
            cell.fill = fill
        for row_idx, ph in enumerate(site_photos, start=2):
            uri = ph.get("uri") or ""
            ph_ws.cell(row=row_idx, column=1, value=row_idx - 1)
            ph_ws.cell(
                row=row_idx,
                column=2,
                value=(ph.get("caption") or "").strip() or f"Photo {row_idx-1}",
            )
            ph_ws.cell(
                row=row_idx,
                column=3,
                value="base64 data URI" if uri.startswith("data:") else "URL",
            )
            ph_ws.cell(row=row_idx, column=4, value=len(uri))
        ph_ws.column_dimensions["A"].width = 4
        ph_ws.column_dimensions["B"].width = 60
        ph_ws.column_dimensions["C"].width = 22
        ph_ws.column_dimensions["D"].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe = _safe_filename(chantier.get("client_name") or chantier_id)
    return StreamingResponse(
        buf,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": (
                f'attachment; filename="MesureChassis_{safe}.xlsx"'
            )
        },
    )



# ═════════════════════════════════════════════════════════════════════════════
# 🆕 V3 — Export ERP universel (CSV + XML générique menuiserie)
#
# Format conçu pour être importé MANUELLEMENT dans n'importe quel ERP de
# menuiserie (Elcia, Ramasoft, BatiPro, etc.) en attendant des specs API
# précises de la part des éditeurs.
#
# Une ligne / un élément par MESURE. Champs prioritaires : référence,
# largeur, hauteur, forme, options. Les cotes spécifiques aux formes
# complexes (trapèze, arc, polygone…) sont sérialisées dans `details`.
# ═════════════════════════════════════════════════════════════════════════════


def _erp_row(chantier: dict, m: dict) -> dict:
    """Construit la représentation ERP normalisée d'une mesure."""
    opts = m.get("options") or {}
    shape = opts.get("shape") or m.get("block_type") or ""
    # Détails libres pour les ERPs qui acceptent un champ "notes" (clé=valeur)
    detail_pairs: list[str] = []
    for key in (
        "width_top",
        "width_bottom",
        "height_left",
        "height_right",
        "trap_height_left",
        "trap_height_right",
        "oeil_diameter",
        "arch_h1_appui",
        "arch_h2_total",
        "angle90_cut_width",
        "angle90_cut_height",
        "polygon_edge_count",
        "polygon_edge_length",
        "polygon_angle_deg",
        "feuillure_left_mm",
        "feuillure_right_mm",
        "feuillure_top_mm",
        "floor_reserve",
    ):
        val = opts.get(key)
        if val not in (None, "", 0, "0"):
            detail_pairs.append(f"{key}={val}")
    return {
        "project_ref": chantier.get("client_name") or "",
        "project_id": chantier.get("id") or "",
        "address": chantier.get("address") or "",
        "postal_code": chantier.get("postal_code") or "",
        "city": chantier.get("city") or "",
        "country": chantier.get("country") or "",
        "item_ref": m.get("label") or m.get("id") or "",
        "item_id": m.get("id") or "",
        "shape": shape,
        "block_type": m.get("block_type") or "",
        "width_mm": m.get("bay_width") or "",
        "height_mm": m.get("bay_height") or "",
        "diagonal_1_mm": m.get("bay_diagonal_1") or m.get("bay_diagonal") or "",
        "diagonal_2_mm": m.get("bay_diagonal_2") or m.get("bay_diagonal") or "",
        "renovation": "yes" if m.get("renovation_mode") else "no",
        "masonry_type": opts.get("masonry_type") or "",
        "wall_thickness_mm": opts.get("gros_oeuvre_mm") or "",
        "insulation_mode": opts.get("insulation_mode") or "",
        "iti_thickness_mm": opts.get("iti_thickness_mm") or "",
        "ite_insul_thickness_mm": opts.get("ite_insul_thickness_mm") or "",
        "details": ";".join(detail_pairs),
        "measured_at": m.get("created_at") or "",
        "measured_by": m.get("created_by_name") or "",
    }


_ERP_FIELDS = [
    "project_ref", "project_id", "address", "postal_code", "city", "country",
    "item_ref", "item_id", "shape", "block_type",
    "width_mm", "height_mm", "diagonal_1_mm", "diagonal_2_mm",
    "renovation", "masonry_type", "wall_thickness_mm",
    "insulation_mode", "iti_thickness_mm", "ite_insul_thickness_mm",
    "details", "measured_at", "measured_by",
]


@router.get("/chantiers/{chantier_id}/export-erp.csv")
async def export_erp_csv(
    chantier_id: str, user=Depends(restrict_advanced_exports)
):
    """Export ERP générique au format CSV (séparateur point-virgule, UTF-8 BOM).

    Le BOM (Byte Order Mark) permet à Excel d'ouvrir le fichier directement
    en UTF-8 sans corrompre les accents (problème historique sur Windows).
    Compatible avec tous les ERPs menuiserie supportant les imports CSV.
    """
    chantier = await db.chantiers.find_one(
        {"id": chantier_id, "company_id": user.get("company_id", "default")},
        {"_id": 0},
    )
    if not chantier:
        raise HTTPException(404, "Chantier introuvable")
    mesures = (
        await db.mesures.find({"chantier_id": chantier_id}, {"_id": 0})
        .sort("created_at", 1)
        .to_list(500)
    )

    buf = io.StringIO()
    # BOM UTF-8 pour Excel
    buf.write("\ufeff")
    writer = csv.DictWriter(
        buf,
        fieldnames=_ERP_FIELDS,
        delimiter=";",
        quoting=csv.QUOTE_MINIMAL,
        extrasaction="ignore",
    )
    writer.writeheader()
    for m in mesures:
        writer.writerow(_erp_row(chantier, m))

    safe = _safe_filename(chantier.get("client_name") or chantier_id)
    return Response(
        content=buf.getvalue().encode("utf-8"),
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": (
                f'attachment; filename="ERP_{safe}.csv"'
            )
        },
    )


@router.get("/chantiers/{chantier_id}/export-erp.xml")
async def export_erp_xml(
    chantier_id: str, user=Depends(restrict_advanced_exports)
):
    """Export ERP générique au format XML.

    Structure : <ERPExport version="1.0"><Project>…</Project></ERPExport>.
    Pensé pour l'import dans Elcia / Ramasoft / autres ERPs supportant
    XML générique. Aucun XSD propriétaire utilisé — l'utilisateur
    pourra mapper les champs manuellement.
    """
    chantier = await db.chantiers.find_one(
        {"id": chantier_id, "company_id": user.get("company_id", "default")},
        {"_id": 0},
    )
    if not chantier:
        raise HTTPException(404, "Chantier introuvable")
    mesures = (
        await db.mesures.find({"chantier_id": chantier_id}, {"_id": 0})
        .sort("created_at", 1)
        .to_list(500)
    )

    def x(v: Any) -> str:
        return xml_escape(str(v)) if v not in (None, "") else ""

    lines: list[str] = []
    lines.append('<?xml version="1.0" encoding="UTF-8"?>')
    lines.append(
        '<ERPExport version="1.0" '
        'generator="MesureChassis" '
        f'generatedAt="{datetime.now(timezone.utc).isoformat()}">'
    )
    lines.append("  <Project>")
    lines.append(f"    <Reference>{x(chantier.get('client_name'))}</Reference>")
    lines.append(f"    <Id>{x(chantier.get('id'))}</Id>")
    lines.append(f"    <Address>{x(chantier.get('address'))}</Address>")
    lines.append(f"    <PostalCode>{x(chantier.get('postal_code'))}</PostalCode>")
    lines.append(f"    <City>{x(chantier.get('city'))}</City>")
    lines.append(f"    <Country>{x(chantier.get('country'))}</Country>")
    lines.append(f"    <Status>{x(chantier.get('status'))}</Status>")
    lines.append(f"    <ItemCount>{len(mesures)}</ItemCount>")
    lines.append("    <Items>")
    for m in mesures:
        row = _erp_row(chantier, m)
        lines.append("      <Item>")
        lines.append(f"        <Reference>{x(row['item_ref'])}</Reference>")
        lines.append(f"        <Id>{x(row['item_id'])}</Id>")
        lines.append(f"        <Shape>{x(row['shape'])}</Shape>")
        lines.append(f"        <BlockType>{x(row['block_type'])}</BlockType>")
        lines.append("        <Dimensions>")
        lines.append(f"          <Width unit=\"mm\">{x(row['width_mm'])}</Width>")
        lines.append(f"          <Height unit=\"mm\">{x(row['height_mm'])}</Height>")
        lines.append(
            f"          <Diagonal1 unit=\"mm\">{x(row['diagonal_1_mm'])}"
            "</Diagonal1>"
        )
        lines.append(
            f"          <Diagonal2 unit=\"mm\">{x(row['diagonal_2_mm'])}"
            "</Diagonal2>"
        )
        lines.append("        </Dimensions>")
        lines.append("        <Wall>")
        lines.append(f"          <Masonry>{x(row['masonry_type'])}</Masonry>")
        lines.append(
            f"          <Thickness unit=\"mm\">{x(row['wall_thickness_mm'])}"
            "</Thickness>"
        )
        lines.append(
            f"          <InsulationMode>{x(row['insulation_mode'])}"
            "</InsulationMode>"
        )
        lines.append(
            f"          <ItiThickness unit=\"mm\">{x(row['iti_thickness_mm'])}"
            "</ItiThickness>"
        )
        lines.append(
            f"          <IteInsulThickness unit=\"mm\">"
            f"{x(row['ite_insul_thickness_mm'])}</IteInsulThickness>"
        )
        lines.append("        </Wall>")
        lines.append(f"        <Renovation>{x(row['renovation'])}</Renovation>")
        lines.append(f"        <Details>{x(row['details'])}</Details>")
        lines.append(f"        <MeasuredAt>{x(row['measured_at'])}</MeasuredAt>")
        lines.append(f"        <MeasuredBy>{x(row['measured_by'])}</MeasuredBy>")
        lines.append("      </Item>")
    lines.append("    </Items>")
    lines.append("  </Project>")
    lines.append("</ERPExport>")

    safe = _safe_filename(chantier.get("client_name") or chantier_id)
    return Response(
        content="\n".join(lines).encode("utf-8"),
        media_type="application/xml; charset=utf-8",
        headers={
            "Content-Disposition": (
                f'attachment; filename="ERP_{safe}.xml"'
            )
        },
    )

