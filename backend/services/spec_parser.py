"""Service de parsing intelligent de cahiers des charges (PDF/Excel/Image).

Objectif :
    Lire un document fourni par l'artisan (cahier des charges client,
    plan d'architecte, tableau Excel…) et en extraire automatiquement
    la liste des châssis à mesurer (type, dimensions, quantité, notes).

Choix techniques :
    • Gemini 2.5 Flash via emergentintegrations.llm.chat (vision native,
      économique, supporte PDF + images natifs grâce à
      `FileContentWithMimeType`).
    • Pour Excel : openpyxl pour lire les cellules, puis on convertit
      en texte tabulé et on demande à Gemini d'extraire les items.
    • L'IA doit renvoyer un JSON strict — on parse + on valide. En cas
      d'échec, on retourne une liste vide et un message d'erreur.
"""
from __future__ import annotations

import json
import logging
import os
import re
import tempfile
from typing import List, Optional

from dotenv import load_dotenv

load_dotenv()
logger = logging.getLogger("mesurechassis.spec_parser")

EMERGENT_LLM_KEY = os.getenv("EMERGENT_LLM_KEY", "")

# Gemini 2.5 Flash : vision multimodale native, rapide, économique.
# Si jamais on bascule sur Pro plus tard, il suffit de changer cette
# constante (l'API LlmChat est identique).
GEMINI_PROVIDER = "gemini"
GEMINI_MODEL = "gemini-2.5-flash"

# Types de châssis supportés dans MesureChâssis (cf. db.VALID_BLOCK_TYPES).
# On force Gemini à choisir parmi ces valeurs pour garantir la cohérence
# avec le reste de l'app (création directe de mesures sans re-mapping).
ALLOWED_BLOCK_TYPES = ["standard", "coulissant", "porte", "trapeze"]

# Prompt système : on est très directif sur le format de sortie.
# Le but est d'obtenir un JSON parsable sans surcouche.
SYSTEM_PROMPT = """Tu es un expert en analyse de cahiers des charges de menuiserie (fenêtres, portes, châssis, volets).

Tu reçois un document (PDF, photo, tableur ou texte) décrivant des ouvertures à installer chez un client.

Ta mission : extraire de façon EXHAUSTIVE la liste des châssis AVEC TOUTES leurs spécifications techniques, ET les exigences globales du projet. AUCUNE donnée technique ne doit être perdue.

# RÈGLES STRICTES DE SORTIE

Réponds UNIQUEMENT avec un objet JSON valide (pas de markdown, pas de commentaire, pas de texte avant/après).

Format :
{
  "items": [
    {
      "label": "string court (ex: 'Fenêtre salon', 'Porte d''entrée')",
      "reference": "repère/plan si présent (ex: 'F1', 'P3', 'M12'), sinon ''",
      "block_type": "standard" | "coulissant" | "porte" | "trapeze",
      "width_mm": entier (largeur mm),
      "height_mm": entier (hauteur mm),
      "quantity": entier (nb d'exemplaires identiques, défaut 1),
      "location": "pièce/étage (ex: 'Cuisine RDC', 'Chambre 1 étage'), sinon ''",
      "material": "matériau (PVC / Aluminium / Bois / Mixte bois-alu), sinon ''",
      "color_ral": "couleur + code RAL si donné (ex: 'Gris anthracite RAL 7016', 'Blanc'), sinon ''",
      "glazing": "type de vitrage (ex: 'Double vitrage 4/16/4', 'Triple', 'Acoustique', 'Sécurit feuilleté 44.2', 'Dépoli'), sinon ''",
      "uw": "coefficient thermique Uw du châssis si donné (ex: '1.3 W/m²K'), sinon ''",
      "ug": "coefficient Ug du vitrage si donné (ex: '1.1'), sinon ''",
      "rw": "affaiblissement acoustique Rw si donné (ex: '38 dB'), sinon ''",
      "opening_type": "type d'ouverture (ex: 'Oscillo-battant', 'Battant', 'Fixe', 'Coulissant', 'Soufflet', 'À galandage'), sinon ''",
      "opening_direction": "sens/main d'ouverture (ex: 'Tirant gauche', 'Poussant droit', '2 vantaux'), sinon ''",
      "security": "exigence sécurité/anti-effraction (ex: 'RC2', 'RC3', 'Vitrage retardateur'), sinon ''",
      "hardware": "quincaillerie (ex: 'Poignée Hoppe alu', 'Serrure 3 points', 'Paumelles renforcées'), sinon ''",
      "accessories": "accessoires (ex: 'Volet roulant intégré', 'Moustiquaire', 'Seuil PMR', 'Grille de ventilation'), sinon ''",
      "notes": "toute autre info utile au menuisier",
      "extra": "TOUTE exigence supplémentaire du document qui n'entre dans AUCUN champ ci-dessus, recopiée mot pour mot. NE RIEN OMETTRE."
    }
  ],
  "project_specs": {
    "client": "nom du client / maître d'ouvrage si présent, sinon ''",
    "address": "adresse du chantier si présente, sinon ''",
    "general_norms": "normes/labels applicables à tout le projet (ex: 'CE', 'NF', 'Acotherm', 'CEKAL', 'DTU 36.5', 'RGE'), sinon ''",
    "deadlines": "dates/délais d'exécution, livraison, pose si présents (ex: 'Pose avant le 15/09/2026', 'Délai 6 semaines'), sinon ''",
    "warranty": "garanties demandées (ex: 'Décennale', 'Garantie 10 ans profilés'), sinon ''",
    "payment": "modalités de paiement si présentes (ex: '30% acompte, solde à la pose'), sinon ''",
    "other": "TOUTE autre exigence globale du document non capturée ailleurs, recopiée fidèlement. NE RIEN OMETTRE."
  },
  "summary": "1 phrase résumant le contenu du document"
}

# RÈGLES MÉTIER

1. **block_type** : EXACTEMENT une de ces 4 valeurs :
   - "porte" : porte d'entrée, porte-fenêtre, porte de service, porte de garage
   - "coulissant" : baie coulissante, châssis coulissant, oscillo-coulissant, galandage
   - "trapeze" : forme trapézoïdale (côtés non parallèles)
   - "standard" : tout le reste (fenêtre classique, châssis fixe, oscillo-battant…)

2. **Dimensions** : toujours en MILLIMÈTRES (entier). Convertis cm/m. Dimension manquante → 0.

3. **Quantity** : "3 fenêtres identiques 1m x 1m20" → UN item quantity=3, width=1000, height=1200. NE répète PAS.

4. **EXHAUSTIVITÉ (CRUCIAL)** : le menuisier doit retrouver dans l'app CHAQUE exigence du cahier des charges. Si une info ne rentre pas dans un champ précis, mets-la dans "extra" (par item) ou "project_specs.other" (global). Recopie fidèlement. NE RÉSUME PAS au point de perdre une exigence (vitrage, coef, norme, RAL, quincaillerie, délai…).

5. **JAMAIS d'inventions** : dimension incertaine → 0 ; type incertain → "standard" ; champ absent → "". Ne devine pas un RAL ou un coefficient non écrit.

6. Aucun châssis détecté → {"items": [], "project_specs": {...champs vides...}, "summary": "Aucun châssis détecté dans ce document."}

7. Sois INDULGENT sur le format d'entrée (tableaux mal alignés, photos floues, plans manuscrits) — fais de ton mieux."""


def _parse_json_response(raw: str) -> dict:
    """Extrait le JSON de la réponse IA (tolère les balises markdown)."""
    if not raw:
        return {"items": [], "summary": "Réponse vide de l'IA."}
    # Strip markdown code fences si présent
    text = raw.strip()
    fence = re.search(r"```(?:json)?\s*(\{.*?\})\s*```", text, flags=re.DOTALL)
    if fence:
        text = fence.group(1)
    else:
        # Cherche le premier { jusqu'au dernier } (au cas où l'IA ajoute du texte)
        first = text.find("{")
        last = text.rfind("}")
        if first != -1 and last != -1 and last > first:
            text = text[first : last + 1]
    try:
        return json.loads(text)
    except json.JSONDecodeError as e:
        logger.warning("JSON IA invalide : %s — payload=%r", e, raw[:500])
        return {"items": [], "summary": f"Erreur parsing IA: {e}"}


# Champs texte détaillés capturés par ouverture (extraction exhaustive CDC).
_ITEM_TEXT_FIELDS = [
    "reference", "location", "material", "color_ral", "glazing", "uw", "ug",
    "rw", "opening_type", "opening_direction", "security", "hardware",
    "accessories", "extra",
]
# Champs des exigences globales du projet.
_PROJECT_FIELDS = [
    "client", "address", "general_norms", "deadlines", "warranty",
    "payment", "other",
]


def _validate_and_clean_items(items_raw: list) -> List[dict]:
    """Nettoie + valide la liste d'items extraite par l'IA.

    Préserve TOUS les champs techniques détaillés (vitrage, Uw/Rw, RAL,
    quincaillerie, normes…) — objectif : ne perdre aucune exigence du CDC.
    """
    cleaned: List[dict] = []
    if not isinstance(items_raw, list):
        return cleaned
    for raw in items_raw:
        if not isinstance(raw, dict):
            continue
        block_type = str(raw.get("block_type") or "standard").strip().lower()
        if block_type not in ALLOWED_BLOCK_TYPES:
            block_type = "standard"
        label = str(raw.get("label") or "").strip()[:80]
        if not label:
            continue
        try:
            width = int(float(raw.get("width_mm") or 0))
            height = int(float(raw.get("height_mm") or 0))
        except (TypeError, ValueError):
            width = height = 0
        try:
            quantity = max(1, int(raw.get("quantity") or 1))
        except (TypeError, ValueError):
            quantity = 1
        item = {
            "label": label,
            "block_type": block_type,
            "width_mm": max(0, width),
            "height_mm": max(0, height),
            "quantity": quantity,
            "notes": str(raw.get("notes") or "").strip()[:1000],
        }
        # Champs techniques détaillés (chaîne, tronqués mais jamais perdus).
        for f in _ITEM_TEXT_FIELDS:
            item[f] = str(raw.get(f) or "").strip()[:600]
        cleaned.append(item)
    return cleaned


def _clean_project_specs(raw: Optional[dict]) -> dict:
    """Nettoie les exigences globales du projet (délais, normes, garanties…)."""
    out = {f: "" for f in _PROJECT_FIELDS}
    if isinstance(raw, dict):
        for f in _PROJECT_FIELDS:
            out[f] = str(raw.get(f) or "").strip()[:1500]
    return out


async def _call_gemini(session_id: str, user_text: str, file_path: Optional[str] = None, mime: Optional[str] = None) -> str:
    """Appelle Gemini 2.5 Flash via emergentintegrations.

    Si `file_path` est fourni → attache le fichier (PDF/Image).
    Sinon → envoie uniquement le texte.
    """
    if not EMERGENT_LLM_KEY:
        raise RuntimeError("EMERGENT_LLM_KEY manquante — impossible d'appeler l'IA.")

    # Import local pour limiter le boot time
    from emergentintegrations.llm.chat import (  # type: ignore
        FileContentWithMimeType,
        LlmChat,
        UserMessage,
    )

    chat = LlmChat(
        api_key=EMERGENT_LLM_KEY,
        session_id=session_id,
        system_message=SYSTEM_PROMPT,
    ).with_model(GEMINI_PROVIDER, GEMINI_MODEL)

    if file_path and mime:
        file_content = FileContentWithMimeType(
            file_path=file_path,
            mime_type=mime,
        )
        msg = UserMessage(text=user_text, file_contents=[file_content])
    else:
        msg = UserMessage(text=user_text)
    reply = await chat.send_message(msg)
    return reply


# ════════════════════════════════════════════════════════════════════════
# Public API
# ════════════════════════════════════════════════════════════════════════
async def parse_pdf(pdf_bytes: bytes, session_id: str) -> dict:
    """Extrait les châssis d'un PDF (cahier des charges, plan, devis client)."""
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
        tmp.write(pdf_bytes)
        tmp_path = tmp.name
    try:
        prompt = (
            "Voici un cahier des charges en PDF. Extrais la liste des "
            "châssis à mesurer au format JSON strict (voir consignes système)."
        )
        raw = await _call_gemini(session_id, prompt, tmp_path, "application/pdf")
    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
    parsed = _parse_json_response(raw)
    return {
        "items": _validate_and_clean_items(parsed.get("items") or []),
        "project_specs": _clean_project_specs(parsed.get("project_specs")),
        "summary": str(parsed.get("summary") or "")[:300],
        "raw_response": raw[:2000],
    }


async def parse_image(img_bytes: bytes, mime: str, session_id: str) -> dict:
    """Extrait les châssis d'une photo (plan papier, page de cahier scannée)."""
    # Normalise le mime
    if mime not in {"image/png", "image/jpeg", "image/jpg", "image/webp"}:
        mime = "image/jpeg"
    suffix = {
        "image/png": ".png",
        "image/jpeg": ".jpg",
        "image/jpg": ".jpg",
        "image/webp": ".webp",
    }.get(mime, ".jpg")
    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
        tmp.write(img_bytes)
        tmp_path = tmp.name
    try:
        prompt = (
            "Voici une photo d'un cahier des charges ou d'un plan de menuiserie. "
            "Extrais la liste des châssis à mesurer au format JSON strict."
        )
        raw = await _call_gemini(session_id, prompt, tmp_path, mime)
    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
    parsed = _parse_json_response(raw)
    return {
        "items": _validate_and_clean_items(parsed.get("items") or []),
        "project_specs": _clean_project_specs(parsed.get("project_specs")),
        "summary": str(parsed.get("summary") or "")[:300],
        "raw_response": raw[:2000],
    }


def _excel_to_text(xlsx_bytes: bytes) -> str:
    """Convertit un .xlsx en texte tabulé (lisible par l'IA)."""
    from openpyxl import load_workbook  # type: ignore

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(xlsx_bytes)
        tmp_path = tmp.name
    try:
        wb = load_workbook(tmp_path, data_only=True)
        chunks: List[str] = []
        # Limite à 5 feuilles max pour éviter les tokens explosifs
        for sheet in wb.worksheets[:5]:
            chunks.append(f"\n=== Feuille : {sheet.title} ===")
            row_count = 0
            for row in sheet.iter_rows(values_only=True):
                # Skip rows entièrement vides
                if not any(cell is not None and str(cell).strip() for cell in row):
                    continue
                line = " | ".join(
                    "" if c is None else str(c).strip() for c in row
                )
                chunks.append(line)
                row_count += 1
                # Limite à 500 lignes par feuille
                if row_count >= 500:
                    chunks.append("[... feuille tronquée à 500 lignes ...]")
                    break
        return "\n".join(chunks)
    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass


async def parse_excel(xlsx_bytes: bytes, session_id: str) -> dict:
    """Extrait les châssis d'un fichier Excel (.xlsx)."""
    try:
        text = _excel_to_text(xlsx_bytes)
    except Exception as e:  # noqa: BLE001
        logger.exception("Excel illisible")
        return {
            "items": [],
            "summary": f"Excel illisible : {type(e).__name__}",
            "raw_response": "",
        }
    if not text.strip():
        return {
            "items": [],
            "summary": "Fichier Excel vide.",
            "raw_response": "",
        }
    prompt = (
        "Voici le contenu textuel d'un fichier Excel (cahier des charges). "
        "Extrais la liste des châssis à mesurer au format JSON strict :\n\n"
        + text[:30000]  # garde-fou sur la taille
    )
    raw = await _call_gemini(session_id, prompt, None, None)
    parsed = _parse_json_response(raw)
    return {
        "items": _validate_and_clean_items(parsed.get("items") or []),
        "project_specs": _clean_project_specs(parsed.get("project_specs")),
        "summary": str(parsed.get("summary") or "")[:300],
        "raw_response": raw[:2000],
    }
